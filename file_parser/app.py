from flask import Flask, jsonify,request
from flask_sqlalchemy import SQLAlchemy
from flask_marshmallow import Marshmallow
from openpyxl import load_workbook

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///todo.db'

db = SQLAlchemy(app)
ma = Marshmallow(app)

class student(db.Model):
    roll_no = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100))
    percentage = db.Column(db.Float)
    branch = db.Column(db.String(50))
 
    def __init__(self,roll_no,name, percentage,branch):
        self.roll_no = roll_no
        self.name = name
        self.percentage = percentage
        self.branch = branch
        
class PostSchema(ma.Schema):
    class Meta:
        fields = ("roll_no", "name", "percentage","branch")
        
post_schema = PostSchema()
posts_schema = PostSchema(many=True)
    
@app.route('/get', methods= ['POST'])
def add_data():
    if request.method == 'POST':
        data = request.files['abdul']
        abdul =load_workbook(data)
        abdul1=abdul.active
        for i in abdul1.iter_rows(min_row=2,values_only = True):
            data1 = student(roll_no=i[0], name=i[1],percentage=i[2],branch=i[3])
            db.session.add(data1)
        db.session.commit()
    return "msg:data retrieve"

@app.route('/retrieve', methods = ['GET'])
def get_post():
    all_posts = student.query.all()
    result = posts_schema.dumps(all_posts)
    
    return jsonify(result)

@app.route('/get_details/<int:roll_no>', methods = ['GET'])
def get_details(roll_no):
    # print("------------------------------------------------>>>>>>>>>>>>>>>>>>>>>>ansjdsbskjcb")
    post = student.query.filter_by(roll_no=roll_no).first()
    result = post_schema.dumps(post)
    
    return jsonify(result)
    
    # return pjsonify(post)

@app.route('/post_delete/<roll_no>/', methods = ['DELETE'])
def post_delete(roll_no):
    post = student.query.get(roll_no)
    db.session.delete(post)
    db.session.commit()
    print(post)
    return post_schema.jsonify(post)

@app.route('/post_updates/<roll_no>/', methods = ['PUT'])
def post_update(roll_no):
    post = student.query.get(roll_no)
    # roll_no = request.json['roll_no']
    name = request.json['name']
    percentage = request.json['percentage']
    branch = request.json['branch']
    
    # post.roll_no = roll_no
    
    post.name = name
    post.percentage = percentage
    post.branch = branch
    
    # db.session.add(post)
    db.session.commit()
    return post_schema.jsonify(post)

with app.app_context():
    db.create_all()
    
if __name__ == "__main__":
     app.run(debug=True ,port=8000,use_reloader=False)
    